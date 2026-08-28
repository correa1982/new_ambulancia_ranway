import io
from flask import render_template, request, session, redirect, url_for, flash, send_file, jsonify
from utils import login_required
from db import get_db
import openpyxl

def register_routes(app):
    @app.route("/personal_operativo", methods=["GET"])
    @login_required
    def personal_operativo_panel():
        if not session.get("usuario") or (not session["usuario"].get("permiso_programacion_operativa") and str(session["usuario"].get("id")) != '1'):
            flash("No tienes permiso.", "error")
            return redirect(url_for("dashboard"))
            
        conn = get_db()
        personal = conn.execute("SELECT * FROM personal_operativo ORDER BY id DESC").fetchall()
        conn.close()
        
        return render_template("personal_operativo.html", personal=personal, usuario=session["usuario"])

    @app.route("/personal_operativo/guardar", methods=["POST"])
    @login_required
    def personal_operativo_guardar():
        if not session.get("usuario") or (not session["usuario"].get("permiso_programacion_operativa") and str(session["usuario"].get("id")) != '1'):
            return jsonify({"success": False, "error": "No autorizado"})
            
        data = request.get_json()
        pid = data.get("id")
        cedula = data.get("cedula", "").strip()
        nombres = data.get("nombres", "").strip()
        apellidos = data.get("apellidos", "").strip()
        codigo_fecha = data.get("codigo_fecha", "").strip()
        registro = data.get("registro", "").strip()
        perfiles = data.get("perfiles", [])
        
        if not all([cedula, nombres, apellidos]):
            return jsonify({"success": False, "error": "Cédula, Nombres y Apellidos son requeridos"})
            
        perfiles_str = ",".join(perfiles) if isinstance(perfiles, list) else str(perfiles)
        
        conn = get_db()
        try:
            if pid:
                conn.execute("""
                    UPDATE personal_operativo
                    SET cedula = ?, nombres = ?, apellidos = ?, codigo_fecha = ?, registro = ?, perfiles = ?
                    WHERE id = ?
                """, (cedula, nombres, apellidos, codigo_fecha, registro, perfiles_str, pid))
            else:
                conn.execute("""
                    INSERT INTO personal_operativo (cedula, nombres, apellidos, codigo_fecha, registro, perfiles, registrado_por)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (cedula, nombres, apellidos, codigo_fecha, registro, perfiles_str, session["usuario"]["id"]))
            conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            conn.rollback()
            return jsonify({"success": False, "error": str(e)})
        finally:
            conn.close()

    @app.route("/personal_operativo/eliminar", methods=["POST"])
    @login_required
    def personal_operativo_eliminar():
        if not session.get("usuario") or (not session["usuario"].get("permiso_programacion_operativa") and str(session["usuario"].get("id")) != '1'):
            return jsonify({"success": False, "error": "No autorizado"})
            
        data = request.get_json()
        pid = data.get("id")
        
        if not pid:
            return jsonify({"success": False, "error": "ID requerido"})
            
        conn = get_db()
        try:
            conn.execute("DELETE FROM personal_operativo WHERE id = ?", (pid,))
            conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            conn.rollback()
            return jsonify({"success": False, "error": str(e)})
        finally:
            conn.close()

    @app.route("/personal_operativo/importar", methods=["POST"])
    @login_required
    def personal_operativo_importar():
        if not session.get("usuario") or (not session["usuario"].get("permiso_programacion_operativa") and str(session["usuario"].get("id")) != '1'):
            flash("No tienes permiso.", "error")
            return redirect(url_for("personal_operativo_panel"))
            
        if 'archivo_excel' not in request.files:
            flash("No se subió ningún archivo", "error")
            return redirect(url_for("personal_operativo_panel"))
            
        file = request.files['archivo_excel']
        if file.filename == '':
            flash("No se seleccionó ningún archivo", "error")
            return redirect(url_for("personal_operativo_panel"))
            
        if not file.filename.endswith('.xlsx'):
            flash("El archivo debe ser un Excel (.xlsx)", "error")
            return redirect(url_for("personal_operativo_panel"))
            
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            # Buscar índices de columnas
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    val = str(cell.value).strip().lower()
                    if 'cedula' in val or 'cédula' in val or 'identificación' in val or 'identificacion' in val:
                        headers['cedula'] = col_idx
                    elif 'nombres' in val or 'nombre' in val:
                        headers['nombres'] = col_idx
                    elif 'apellidos' in val or 'apellido' in val:
                        headers['apellidos'] = col_idx
                    elif 'codigo' in val or 'código' in val or 'fecha' in val:
                        headers['codigo_fecha'] = col_idx
                    elif 'registro' in val:
                        headers['registro'] = col_idx
                    elif 'perfil' in val or 'perfiles' in val:
                        headers['perfiles'] = col_idx
            
            if not all(k in headers for k in ['cedula', 'nombres', 'apellidos']):
                flash("El Excel debe contener las columnas: Cédula, Nombres, Apellidos", "error")
                return redirect(url_for("personal_operativo_panel"))
            
            conn = get_db()
            registros_exitosos = 0
            
            # Upsert
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                cedula = row[headers['cedula'] - 1].value
                nombres = row[headers['nombres'] - 1].value
                apellidos = row[headers['apellidos'] - 1].value
                
                if not cedula or not nombres or not apellidos:
                    continue
                    
                cedula = str(cedula).strip()
                nombres = str(nombres).strip()
                apellidos = str(apellidos).strip()
                
                codigo_fecha = ""
                if 'codigo_fecha' in headers:
                    cf_val = row[headers['codigo_fecha'] - 1].value
                    if cf_val:
                        if hasattr(cf_val, 'strftime'):
                            codigo_fecha = cf_val.strftime('%d/%m/%Y')
                        else:
                            codigo_fecha = str(cf_val).strip()
                        
                registro = ""
                if 'registro' in headers:
                    r_val = row[headers['registro'] - 1].value
                    if r_val:
                        registro = str(r_val).strip()
                        
                perfiles = ""
                if 'perfiles' in headers:
                    p_val = row[headers['perfiles'] - 1].value
                    if p_val:
                        perfiles = str(p_val).strip()
                        
                existente = conn.execute("SELECT id FROM personal_operativo WHERE cedula = ?", (cedula,)).fetchone()
                if existente:
                    conn.execute("""
                        UPDATE personal_operativo 
                        SET nombres=?, apellidos=?, codigo_fecha=?, registro=?, perfiles=? 
                        WHERE cedula=?
                    """, (nombres, apellidos, codigo_fecha, registro, perfiles, cedula))
                else:
                    conn.execute("""
                        INSERT INTO personal_operativo (cedula, nombres, apellidos, codigo_fecha, registro, perfiles, registrado_por)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (cedula, nombres, apellidos, codigo_fecha, registro, perfiles, session["usuario"]["id"]))
                
                registros_exitosos += 1
                
            conn.commit()
            conn.close()
            flash(f"Se procesaron {registros_exitosos} registros correctamente.", "success")
        except Exception as e:
            flash(f"Error al procesar el archivo: {e}", "error")
            
        return redirect(url_for("personal_operativo_panel"))

    @app.route("/personal_operativo/exportar", methods=["GET"])
    @login_required
    def personal_operativo_exportar():
        if not session.get("usuario") or (not session["usuario"].get("permiso_programacion_operativa") and str(session["usuario"].get("id")) != '1'):
            return "No autorizado", 403
            
        conn = get_db()
        personal = conn.execute("SELECT * FROM personal_operativo ORDER BY id DESC").fetchall()
        conn.close()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Personal Operativo"
        
        headers = ['Cédula', 'Nombres', 'Apellidos', 'Código/Fecha', 'Registro', 'Perfiles']
        ws.append(headers)
        
        for p in personal:
            ws.append([
                p.get('cedula', ''),
                p.get('nombres', ''),
                p.get('apellidos', ''),
                p.get('codigo_fecha', ''),
                p.get('registro', ''),
                p.get('perfiles', '')
            ])
            
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        
        return send_file(
            excel_io,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Personal_Operativo.xlsx'
        )
