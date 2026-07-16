import os
import uuid
from datetime import datetime
from flask import Blueprint, request, jsonify, render_template, current_app, send_from_directory, flash, redirect, url_for
from werkzeug.utils import secure_filename
from db import get_db
from utils import login_required, get_user_info

def register_routes(app):
    nomina_bp = Blueprint('nomina', __name__)

    # Ensure upload directory exists
    UPLOAD_FOLDER = os.path.join(app.root_path, 'static', 'uploads', 'nomina')
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    
    # We will only allow Excel extensions for safety
    ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

    def allowed_file(filename):
        return '.' in filename and \
               filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

    @nomina_bp.route('/nomina', methods=['GET'])
    @login_required
    def nomina_index():
        from flask import session
        user_info = session['usuario']
        if user_info.get('rol_real') != 'admin' and 'nomina' not in user_info.get('formularios_acceso', []):
            flash('Acceso no autorizado al módulo de nómina.', 'error')
            return redirect(url_for('dashboard'))
        conn = get_db()
        try:
            cursor = conn.execute("SELECT * FROM nomina ORDER BY fecha_subida DESC")
            archivos = cursor.fetchall()
        except Exception as e:
            current_app.logger.error(f"Error fetching nomina records: {e}")
            archivos = []
        finally:
            conn.close()

        return render_template('nomina.html', user_info=user_info, archivos=archivos)

    @nomina_bp.route('/nomina/upload', methods=['POST'])
    @login_required
    def nomina_upload():
        from flask import session
        user_info = session['usuario']
        if user_info.get('rol_real') != 'admin' and 'nomina' not in user_info.get('formularios_acceso', []):
            flash('Acceso no autorizado al módulo de nómina.', 'error')
            return redirect(url_for('dashboard'))

        if 'file' not in request.files:
            flash('No se seleccionó ningún archivo.', 'error')
            return redirect(url_for('nomina.nomina_index'))
            
        file = request.files['file']
        if file.filename == '':
            flash('El nombre del archivo está vacío.', 'error')
            return redirect(url_for('nomina.nomina_index'))
            
        if file and allowed_file(file.filename):
            original_filename = secure_filename(file.filename)
            # Make the filename unique to prevent overwriting
            unique_filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:6]}_{original_filename}"
            file_path = os.path.join(UPLOAD_FOLDER, unique_filename)
            
            try:
                file.save(file_path)
                
                # Save to database
                from flask import session
                import json
                import openpyxl
                user_info = session['usuario']
                conn = get_db()
                try:
                    cursor = conn.execute("""
                        INSERT INTO nomina (fecha_subida, nombre_archivo, archivo_url, registrado_por, registrado_por_identificacion)
                        VALUES (?, ?, ?, ?, ?)
                    """, (
                        datetime.now(),
                        original_filename,
                        unique_filename,
                        user_info['nombre'],
                        user_info['identificacion']
                    ))
                    nomina_id = cursor.lastrowid
                    
                    # Parse logic
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                    ws = wb.active
                    
                    try:
                        conn.execute("ALTER TABLE nomina ADD COLUMN periodo VARCHAR(255)")
                    except Exception:
                        pass
                        
                    periodo_val = str(ws['U2'].value).strip() if ws['U2'].value else ""
                    conn.execute("UPDATE nomina SET periodo = ? WHERE id = ?", (periodo_val, nomina_id))
                    
                    header_row_idx = None
                    headers = []
                    for idx, row in enumerate(ws.iter_rows(values_only=True)):
                        row_strs = [str(c).strip().upper() for c in row if c is not None]
                        if any('CODIGO' in c for c in row_strs) or any('IDENTIFICACI' in c for c in row_strs):
                            header_row_idx = idx
                            headers = [str(c).strip() if c is not None else f"Col_{i}" for i, c in enumerate(row)]
                            break
                    
                    if header_row_idx is not None:
                        cc_idx, nombres_idx, apellidos_idx, codigo_idx, total_idx = -1, -1, -1, -1, -1
                        for i, h in enumerate(headers):
                            hu = h.upper()
                            if 'IDENTIFICACI' in hu: cc_idx = i
                            elif 'NOMBRES' in hu: nombres_idx = i
                            elif 'APELLIDOS' in hu: apellidos_idx = i
                            elif 'CODIGO' in hu: codigo_idx = i
                            elif 'TOTAL' in hu: total_idx = i
                        
                        for idx, row in enumerate(ws.iter_rows(values_only=True)):
                            if idx <= header_row_idx:
                                continue
                            if not row or cc_idx == -1 or len(row) <= cc_idx or row[cc_idx] is None:
                                continue
                            cc_val = str(row[cc_idx]).strip()
                            if not cc_val or cc_val.lower() == 'none':
                                continue
                                
                            nombres_val = str(row[nombres_idx]).strip() if nombres_idx != -1 and len(row) > nombres_idx and row[nombres_idx] is not None else ""
                            apellidos_val = str(row[apellidos_idx]).strip() if apellidos_idx != -1 and len(row) > apellidos_idx and row[apellidos_idx] is not None else ""
                            codigo_val = str(row[codigo_idx]).strip() if codigo_idx != -1 and len(row) > codigo_idx and row[codigo_idx] is not None else ""
                            total_val = str(row[total_idx]).strip() if total_idx != -1 and len(row) > total_idx and row[total_idx] is not None else ""
                            
                            detalle_dict = {}
                            for i, cell_val in enumerate(row):
                                if i not in (cc_idx, nombres_idx, apellidos_idx, codigo_idx, total_idx) and i < len(headers):
                                    h_name = headers[i]
                                    if cell_val is not None and h_name and not h_name.startswith('Col_'):
                                        # Format floats properly
                                        val_str = str(cell_val)
                                        if isinstance(cell_val, float):
                                            val_str = f"{cell_val:,.2f}"
                                        detalle_dict[h_name] = val_str
                                        
                            detalle_json = json.dumps(detalle_dict, ensure_ascii=False)
                            
                            conn.execute("""
                                INSERT INTO nomina_empleados (nomina_id, identificacion, nombres, apellidos, codigo, valor_total, detalle)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                            """, (nomina_id, cc_val, nombres_val, apellidos_val, codigo_val, total_val, detalle_json))
                            
                    conn.commit()
                    flash('Archivo de nómina subido y procesado exitosamente.', 'success')
                except Exception as db_error:
                    conn.close()
                    # Clean up the file if db insert fails
                    if os.path.exists(file_path):
                        os.remove(file_path)
                    current_app.logger.error(f"Error saving nomina to db: {db_error}")
                    flash('Error al guardar el registro en la base de datos.', 'error')
                finally:
                    conn.close()
                    
            except Exception as save_error:
                current_app.logger.error(f"Error saving nomina file: {save_error}")
                flash('Error al guardar el archivo en el servidor.', 'error')
                
        else:
            flash('Tipo de archivo no permitido. Solo se permiten archivos Excel o CSV (.xlsx, .xls, .csv).', 'error')
            
        return redirect(url_for('nomina.nomina_index'))

    @nomina_bp.route('/nomina/download/<filename>')
    @login_required
    def download_nomina(filename):
        return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)

    @nomina_bp.route('/nomina/delete/<int:id>', methods=['POST'])
    @login_required
    def delete_nomina(id):
        conn = get_db()
        try:
            cursor = conn.execute("SELECT archivo_url FROM nomina WHERE id = ?", (id,))
            record = cursor.fetchone()
            if record:
                file_path = os.path.join(UPLOAD_FOLDER, record['archivo_url'])
                if os.path.exists(file_path):
                    os.remove(file_path)
                    
                conn.execute("DELETE FROM nomina WHERE id = ?", (id,))
                conn.commit()
                flash('Registro eliminado exitosamente.', 'success')
            else:
                flash('Registro no encontrado.', 'error')
        except Exception as e:
            current_app.logger.error(f"Error deleting nomina record: {e}")
            flash('Error al eliminar el registro.', 'error')
        finally:
            conn.close()
            
        return redirect(url_for('nomina.nomina_index'))

    @nomina_bp.route('/nomina/consulta', methods=['GET', 'POST'])
    def nomina_consulta():
        import json
        if request.method == 'POST':
            cedula = request.form.get('cedula', '').strip()
            codigo = request.form.get('codigo', '').strip()
            
            if not cedula or not codigo:
                flash('Debe ingresar la cédula y el código.', 'error')
                return render_template('nomina_consulta.html', resultado=None)
                
            conn = get_db()
            try:
                # Find the latest nomina uploaded
                cursor = conn.execute("SELECT id, fecha_subida, periodo FROM nomina ORDER BY fecha_subida DESC LIMIT 1")
                latest_nomina = cursor.fetchone()
                
                if not latest_nomina:
                    flash('No hay registros de nómina en el sistema.', 'error')
                    return render_template('nomina_consulta.html', resultado=None)
                    
                nomina_id = latest_nomina['id']
                
                # Check for employee in this nomina
                cursor = conn.execute("""
                    SELECT * FROM nomina_empleados 
                    WHERE nomina_id = ? AND identificacion = ? AND codigo = ?
                """, (nomina_id, cedula, codigo))
                empleado = cursor.fetchone()
                
                if empleado:
                    # Parse detalle json
                    if empleado['detalle']:
                        try:
                            empleado['detalle_dict'] = json.loads(empleado['detalle'])
                        except:
                            empleado['detalle_dict'] = {}
                    else:
                        empleado['detalle_dict'] = {}
                        
                    return render_template('nomina_consulta.html', resultado=empleado, fecha_nomina=latest_nomina['fecha_subida'], periodo_nomina=latest_nomina.get('periodo', ''))
                else:
                    flash('Cédula o código incorrectos, o no se encontró en la nómina actual.', 'error')
            except Exception as e:
                current_app.logger.error(f"Error querying nomina: {e}")
                flash('Error de servidor al consultar la nómina.', 'error')
            finally:
                conn.close()
                
        return render_template('nomina_consulta.html', resultado=None)

    app.register_blueprint(nomina_bp)
